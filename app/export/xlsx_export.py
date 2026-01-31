# -*- coding: utf-8 -*-
"""
Экспорт результатов в Excel (.xlsx)
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..models.project import Project
from ..models.calculation import CalculationResult


def export_to_xlsx(project: Project, result: CalculationResult, file_path: str):
    """Экспорт результатов расчёта в Excel"""
    wb = Workbook()

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # === Лист 1: Общие сведения ===
    ws1 = wb.active
    ws1.title = 'Общие сведения'

    ws1['A1'] = 'РАСЧЁТ ТРУДОЁМКОСТИ РАЗРАБОТКИ ПС'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')

    ws1['A3'] = 'Название ПС:'
    ws1['B3'] = project.name
    ws1['A4'] = 'Описание:'
    ws1['B4'] = project.description or '-'
    ws1['A5'] = 'Фонд рабочего времени:'
    ws1['B5'] = f'{project.work_fund} дней/месяц'
    ws1['A6'] = 'Тип ограничения:'
    ws1['B6'] = 'По продолжительности' if project.constraint_type == 'duration' else 'По численности'
    ws1['A7'] = 'Значение ограничения:'
    ws1['B7'] = f'{project.constraint_value} {"месяцев" if project.constraint_type == "duration" else "человек"}'

    ws1['A9'] = 'Дата расчёта:'
    ws1['B9'] = datetime.now().strftime('%d.%m.%Y %H:%M')

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 50

    # === Лист 2: Функции ===
    ws2 = wb.create_sheet('Функции')

    headers = ['Компонент', 'ID функции', 'Название функции', 'Vi', 'ri', 'ki', 'Vm', 'K_slozhn', 'K_sr_razr', 'K_opyt', 'Vk']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    row = 2
    for fr in result.functions_results:
        ws2.cell(row=row, column=1, value=fr.component_name).border = thin_border
        ws2.cell(row=row, column=2, value=fr.function_id).border = thin_border
        ws2.cell(row=row, column=3, value=fr.function_name).border = thin_border
        ws2.cell(row=row, column=4, value=fr.volume_base).border = thin_border
        ws2.cell(row=row, column=5, value=fr.reuse_count).border = thin_border
        ws2.cell(row=row, column=6, value=fr.reuse_coefficient).border = thin_border
        ws2.cell(row=row, column=7, value=fr.volume_corrected).border = thin_border
        ws2.cell(row=row, column=8, value=fr.k_slozhn).border = thin_border
        ws2.cell(row=row, column=9, value=fr.k_sr_razr).border = thin_border
        ws2.cell(row=row, column=10, value=fr.k_opyt).border = thin_border
        ws2.cell(row=row, column=11, value=fr.volume_adjusted).border = thin_border
        row += 1

    # Итого
    ws2.cell(row=row, column=1, value='ИТОГО').font = header_font
    ws2.cell(row=row, column=7, value=sum(fr.volume_corrected for fr in result.functions_results)).font = header_font
    ws2.cell(row=row, column=11, value=result.total_volume).font = header_font

    for col in range(1, 12):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    ws2.column_dimensions['C'].width = 40

    # === Лист 3: Коэффициенты ===
    ws3 = wb.create_sheet('Коэффициенты')

    ws3['A1'] = 'Коэффициенты уровня расчёта'
    ws3['A1'].font = Font(bold=True, size=12)
    ws3.merge_cells('A1:C1')

    coef_headers = ['Коэффициент', 'Значение', 'Числовое значение']
    for col, header in enumerate(coef_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    coeffs = project.coefficients
    coef_data = [
        ('K_n (степень новизны)', coeffs.novelty, result.k_n),
        ('K_nad (надёжность)', coeffs.reliability, result.k_nad),
        ('K_proizv (производительность)', coeffs.performance, result.k_proizv),
        ('K_dokum (документация)', coeffs.documentation, result.k_dokum),
        ('K_teh (технологии)', f'K_str={coeffs.structure}', result.k_teh),
        ('K_or (опыт разработки)', coeffs.dev_experience, result.k_or),
        ('K_sr_srok (влияние сроков)', coeffs.deadline, result.k_sr_srok),
    ]

    for row, (name, value, num) in enumerate(coef_data, 4):
        ws3.cell(row=row, column=1, value=name).border = thin_border
        ws3.cell(row=row, column=2, value=str(value)[:50]).border = thin_border
        ws3.cell(row=row, column=3, value=num).border = thin_border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 20

    # === Лист 4: Подпроцессы ===
    ws4 = wb.create_sheet('Подпроцессы')

    ws4['A1'] = 'Трудоёмкость подпроцессов'
    ws4['A1'].font = Font(bold=True, size=12)
    ws4.merge_cells('A1:E1')

    sp_headers = ['Подпроцесс', 'Коэфф. A', 'Трудоёмкость [чел.-дн.]', 'Численность [чел.]', 'Срок [мес.]']
    for col, header in enumerate(sp_headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    row = 4
    for sp in result.subprocess_results:
        ws4.cell(row=row, column=1, value=sp.name).border = thin_border
        ws4.cell(row=row, column=2, value=sp.base_coefficient).border = thin_border
        ws4.cell(row=row, column=3, value=sp.labor).border = thin_border
        ws4.cell(row=row, column=4, value=sp.staff).border = thin_border
        ws4.cell(row=row, column=5, value=sp.duration).border = thin_border
        row += 1

    # Итого
    for col in range(1, 6):
        ws4.cell(row=row, column=col).font = header_font
        ws4.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws4.cell(row=row, column=col).border = thin_border

    ws4.cell(row=row, column=1, value='ИТОГО')
    ws4.cell(row=row, column=2, value=1.00)
    ws4.cell(row=row, column=3, value=result.total_labor)
    ws4.cell(row=row, column=4, value=result.average_staff)
    ws4.cell(row=row, column=5, value=result.total_duration)

    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 20

    # === Лист 5: Итоги ===
    ws5 = wb.create_sheet('Итоги')

    ws5['A1'] = 'ИТОГОВЫЕ РЕЗУЛЬТАТЫ РАСЧЁТА'
    ws5['A1'].font = Font(bold=True, size=14)
    ws5.merge_cells('A1:C1')

    ws5['A3'] = 'Показатель'
    ws5['B3'] = 'Значение'
    ws5['C3'] = 'Единица измерения'
    for col in range(1, 4):
        ws5.cell(row=3, column=col).font = header_font
        ws5.cell(row=3, column=col).fill = header_fill
        ws5.cell(row=3, column=col).border = thin_border

    results_data = [
        ('Общий объём ПС (V)', result.total_volume, 'строк'),
        ('Базовая трудоёмкость (T_baz)', result.base_labor, 'чел.-дн.'),
        ('Трудоёмкость разработки (T_razr)', result.total_labor, 'чел.-дн.'),
        ('Итоговая трудоёмкость (T_srok)', result.final_labor, 'чел.-дн.'),
        ('Общий срок разработки', result.total_duration, 'месяцев'),
        ('Средняя численность', result.average_staff, 'человек'),
    ]

    for row, (name, value, unit) in enumerate(results_data, 4):
        ws5.cell(row=row, column=1, value=name).border = thin_border
        ws5.cell(row=row, column=2, value=round(value, 2)).border = thin_border
        ws5.cell(row=row, column=3, value=unit).border = thin_border

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 20

    # Формула
    ws5['A12'] = 'Формула расчёта базовой трудоёмкости:'
    ws5['A12'].font = Font(bold=True)
    ws5['A13'] = 'T_baz = A × V^C × K_n × K_nad × K_proizv × K_dokum × K_teh × K_or'
    ws5['A14'] = f'T_baz = {result.A} × {result.total_volume:.2f}^{result.C} × {result.k_n} × {result.k_nad} × {result.k_proizv} × {result.k_dokum} × {result.k_teh:.4f} × {result.k_or}'
    ws5['A15'] = f'T_baz = {result.base_labor:.2f} чел.-дн.'
    ws5['A15'].font = Font(bold=True)

    # Сохранение
    wb.save(file_path)
